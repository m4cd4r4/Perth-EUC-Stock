# Build Room\build_roomv3.6.py
# Author: Macdara O Murchu
# 09.05.24

# Workbook is now selected via a File Explorer dialog
# DONE(?) # Debug: Data validation may be disabled for SANs - Fix: Following had gone missing somehow - "from tkinter import messagebox"
# To-Do: When app is closed, terminal stays loaded
# DONE # Plot scripts still point to old workbook path for data. Change to match main script.

# Create config.py to save workbook apth, so that it can be loaded into the "inventory-levels" scripts

import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
from tkinter import filedialog
from tkinter import messagebox

# Function to save the workbook path to config.py
def save_config(workbook_path):
    with open('config.py', 'w') as config_file:
        config_file.write(f"workbook_path = r'{workbook_path}'\n")


logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)


def run_inventory_script():
    script_path = script_directory / "inventory-levels_4.2v2.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_4.2v2.py' does not exist in the directory.")

def run_build_room_inventory_script():
    script_path = script_directory / "inventory-levels_BRv2.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_BRv2.py' does not exist in the directory.")

def run_combined_rooms_inventory_script():
    script_path = script_directory / "inventory-levels_combinedv1.2.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_combinedv1.2.py' does not exist in the directory.")

def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")


def view_all_sans_log():
    log_window = tk.Toplevel(root)
    log_window.title("SANs In Stock")
    log_window.geometry("600x800")

    # Create a Treeview widget to display the log
    columns = ("SAN Number", "Item", "Timestamp")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="w")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "All SANs" sheet
    if 'All_SANs' in workbook.sheetnames:
        all_sans_sheet = workbook['All_SANs']
        for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
            log_tree.insert('', 'end', values=row)
    else:
        tk.messagebox.showinfo("Info", "All_SANs log is empty.", parent=log_window)


root = ctk.CTk()
root.title("Perth EUC Stock")
root.geometry("675x850")

menu_bar = tk.Menu(root)
plots_menu = tk.Menu(menu_bar, tearoff=0)
plots_menu.add_command(label="Basement 4.2 Inventory", command=run_inventory_script)
plots_menu.add_command(label="Build Room Inventory", command=run_build_room_inventory_script)
plots_menu.add_command(label="Combined Inventory", command=run_combined_rooms_inventory_script)
plots_menu.add_command(label="SANs In Stock", command=view_all_sans_log)
plots_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)
# plots_menu.add_command(label="Headsets In Stock", command=view_headsets_log)
menu_bar.add_cascade(label="Data", menu=plots_menu)
root.config(menu=menu_bar)

script_directory = Path(__file__).parent

# workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'
# if Path(workbook_path).exists():
#     workbook = load_workbook(workbook_path)
# else:
#     workbook = Workbook()
#     workbook.active.title = '4.2_Items'
#     workbook.create_sheet('4.2_Timestamps')
#     workbook.create_sheet('BR_Items')
#     workbook.create_sheet('BR_Timestamps')
#     workbook.create_sheet('All_SANs')
#     workbook['4.2_Items'].append(["Item", "LastCount", "NewCount"])
#     workbook['4.2_Timestamps'].append(["Timestamp", "Item", "Action", "SAN_Number"])
#     workbook['BR_Items'].append(["Item", "LastCount", "NewCount"])
#     workbook['BR_Timestamps'].append(["Timestamp", "Item", "Action", "SAN_Number"])
#     workbook['All_SANs'].append(["SAN_Number", "Item", "Timestamp"])
#     workbook.save(workbook_path)

def get_file_path():
    # Creating a temporary root window for file dialog
    temp_root = tk.Tk()
    temp_root.withdraw()  # Hide the temporary root window
    file_path = filedialog.askopenfilename(
        title="Select a spreadsheet file",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    temp_root.destroy()  # Destroy the temporary root window to clean up
    if not file_path:
        tk.messagebox.showerror("Error", "No file selected. Exiting application.")
        raise SystemExit  # Exit the application if no file is selected
    return file_path

# Get the workbook path from user
workbook_path = get_file_path()
save_config(workbook_path)  # Save the path to the config file immediately after getting it
workbook = load_workbook(workbook_path)

all_sans_sheet = workbook['All_SANs']
sheets = {'original': ('4.2_Items', '4.2_Timestamps'), 'backup': ('BR_Items', 'BR_Timestamps')}
current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12,))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')

class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN number.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()

def is_san_unique(san_number):
    # Adjust the search to account for the 'SAN' prefix properly
    search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
    unique = all(search_string != row[0] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))
    print(f"Checking SAN {search_string}: Unique - {unique}")  # Debug print
    return unique


def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result

def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

frame = ctk.CTkFrame(root)
frame.pack(padx=3, pady=3, fill='both', expand=True)
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=3)

button_width = 25
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_1.pack(side='left', padx=3)
button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_2.pack(side='left', padx=(3, 50))
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_subtract.pack(side='left', padx=3)
entry_value = tk.Entry(entry_frame, font=("Helvetica", 14), justify='center', width=5, validate="key", validatecommand=vcmd)
entry_value.pack(side='left', padx=3)
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_add.pack(side='left', padx=3)
# xlsx_button = ctk.CTkButton(entry_frame, text=".xlsx", command=open_spreadsheet, width=button_width, font=("Helvetica", 14))
# xlsx_button.pack(side='left', padx=3)

def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1

def log_change(item, action, san_number="", timestamp_sheet=None, volume=1):  # Added volume parameter with default value of 1
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        if timestamp_sheet is not None:
            san_number = f"SAN{san_number}" if san_number and not san_number.startswith('SAN') else san_number
            
            # Conditionally modify the action string to include volume for non-SAN items
            if san_number == "":
                action_text = f"{action} {volume}"
            else:
                action_text = action
            
            timestamp_sheet.append([timestamp, item, action_text, san_number])  # Use action_text instead of action
            workbook.save(workbook_path)
            update_log_view()
            logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action_text}, SAN: {san_number}")  # Use action_text
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")

def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        # Adjust the sorting to use the first column (timestamp)
        sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
        row_count = 0
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1


def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        input_value = entry_value.get()
        if input_value.isdigit():
            input_value = int(input_value)
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]
            san_required = any(g in selected_item for g in ["G8", "G9", "G10"])

            entered_sans_count = 0  # To keep track of successfully entered SAN numbers if needed

            if san_required:
                # Loop to input each SAN for SAN-required items
                while entered_sans_count < input_value:
                    san_number = show_san_input()
                    if san_number is None:  # User cancelled the input
                        break  # Keep the already entered SANs
                    # Ensure SAN number has the 'SAN' prefix
                    san_number = "SAN" + san_number if not san_number.startswith("SAN") else san_number
                    
                    if operation == 'add':
                        if is_san_unique(san_number):
                            all_sans_sheet.append([san_number, selected_item, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                            # Log each SAN unique number immediately
                            log_change(selected_item, operation, san_number, timestamp_sheet, volume=1)
                            workbook.save(workbook_path)
                            entered_sans_count += 1
                        else:
                            tk.messagebox.showerror("Error", "Duplicate or already used SAN number.", parent=root)
                    elif operation == 'subtract':
                        # Search for the SAN in all_sans_sheet and remove if found and matches the item
                        for row in all_sans_sheet.iter_rows(min_row=2):
                            if row[0].value == san_number and row[1].value == selected_item:
                                all_sans_sheet.delete_rows(row[0].row)
                                log_change(selected_item, operation, san_number, timestamp_sheet, volume=1)
                                workbook.save(workbook_path)
                                entered_sans_count += 1
                                break
                        else:  # Executed if the loop completes without breaking (SAN not found or doesn't match item)
                            tk.messagebox.showerror("Error", f"SAN number {san_number} does not match the selected item.", parent=root)


            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    row[1].value = row[2].value or 0  # Update LastCount to the current NewCount
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + (input_value if not san_required else entered_sans_count)
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - (input_value if not san_required else entered_sans_count), 0)

            # Log volume for non-SAN items or after all entered SANs
            if (san_required and entered_sans_count > 0) or not san_required:
                volume_to_log = input_value if not san_required else entered_sans_count
                log_change(selected_item, operation, "", timestamp_sheet, volume=volume_to_log)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()

columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column("Item", anchor='w', width=250, stretch=False) # Width of the "Item" column in the treeview. The other columns are default width.
    tree.column("LastCount", anchor='w', width=175, stretch=False)
tree.pack(expand=True, fill="both", padx=3, pady=3)

log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column("Timestamp", anchor='w', width=175, stretch=False)
    log_view.column("Item", anchor='w', width=180, stretch=False)
    log_view.column("Action", anchor='w', width=100, stretch=False)

scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')

# Add the copying functionality here
def add_copy_option(tree):
    def copy_selection():
        selected_item = tree.item(tree.focus())['values']
        item_text = ", ".join(map(str, selected_item))
        root.clipboard_clear()  # Clear the clipboard
        root.clipboard_append(item_text)  # Append new value to the clipboard

    # Create a menu
    context_menu = tk.Menu(tree, tearoff=0)
    context_menu.add_command(label="Copy", command=copy_selection)

    # Function to show the menu
    def show_context_menu(event):
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", show_context_menu)  # Bind right-click event

# Apply the add_copy_option function to the Treeviews
add_copy_option(tree)
add_copy_option(log_view)

root.after(100, update_treeview)
update_log_view()

root.mainloop()

